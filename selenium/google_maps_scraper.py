
import time
import re
import logging
from dataclasses import dataclass, field
from typing import Optional

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


#  CONFIGURATION — edit these values as needed

SEARCH_QUERY   = "restaurants in New York"   # Google Maps search query
MAX_RESULTS    = 60                           # Max number of places to scrape
OUTPUT_FILE    = "nyc_restaurants.xlsx"       # Output Excel filename
SCROLL_PAUSE   = 2.0                          # Seconds to wait between scrolls
PAGE_LOAD_WAIT = 10                           # Max seconds to wait for page elements
HEADLESS       = False                        # Set True to run browser in background


#  LOGGING SETUP

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)



#  DATA MODEL

@dataclass
class Restaurant:
    """Holds all scraped data for a single restaurant."""
    name:         str            = "N/A"
    address:      str            = "N/A"
    phone:        str            = "N/A"
    rating:       Optional[float] = None
    review_count: int            = 0
    website:      str            = "N/A"
    cuisine:      str            = "N/A"
    maps_url:     str            = "N/A"


# ─────────────────────────────────────────────
#  BROWSER SETUP
# ─────────────────────────────────────────────
def build_driver(headless: bool = False) -> webdriver.Chrome:
    """
    Configure and return a Chrome WebDriver instance.
    Adds stealth options to reduce bot-detection risk.
    """
    options = Options()

    if headless:
        options.add_argument("--headless=new")

    # General stealth & stability options
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1400,900")
    options.add_argument("--lang=en-US")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )

    driver = webdriver.Chrome(options=options)
    # Hide webdriver property from JavaScript
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    log.info("Chrome WebDriver started successfully.")
    return driver


# ─────────────────────────────────────────────
#  SCRAPING HELPERS
# ─────────────────────────────────────────────
def safe_find_text(driver, css_selector: str, default: str = "N/A") -> str:
    """
    Safely extract text from a CSS selector.
    Returns `default` if the element is not found.
    """
    try:
        el = driver.find_element(By.CSS_SELECTOR, css_selector)
        return el.text.strip() or default
    except NoSuchElementException:
        return default


def parse_rating(raw: str) -> Optional[float]:
    """Convert a rating string like '4.5' to float, or None if invalid."""
    try:
        return float(raw.replace(",", "."))
    except (ValueError, AttributeError):
        return None


def parse_review_count(raw: str) -> int:
    """
    Extract integer review count from strings like '(1,234)' or '1234 reviews'.
    Returns 0 if parsing fails.
    """
    digits = re.sub(r"[^\d]", "", raw)
    return int(digits) if digits else 0


# ─────────────────────────────────────────────
#  MAIN SCRAPER
# ─────────────────────────────────────────────
def scrape_google_maps(query: str, max_results: int) -> list[Restaurant]:
    """
    Open Google Maps, search for `query`, scroll through results,
    click each listing, and extract restaurant details.

    Returns a list of Restaurant dataclass instances.
    """
    driver = build_driver(headless=HEADLESS)
    wait   = WebDriverWait(driver, PAGE_LOAD_WAIT)
    results: list[Restaurant] = []

    try:
        # ── Step 1: Open Google Maps ──────────────────────────────────────
        log.info("Opening Google Maps …")
        # Force English language via URL parameter
        driver.get("https://www.google.com/maps?hl=en")
        time.sleep(4)  # Wait longer for full page load

        # Dismiss cookie consent if it appears (EU regions)
        try:
            consent_btn = driver.find_element(
                By.XPATH, '//button[contains(., "Accept all") or contains(., "Reject all")]'
            )
            consent_btn.click()
            time.sleep(1)
        except NoSuchElementException:
            pass

        # ── Step 2: Enter search query ────────────────────────────────────
        log.info(f"Searching for: '{query}'")

        # Try multiple selectors for the search box (Google changes these sometimes)
        search_box = None
        selectors = [
            (By.ID, "searchboxinput"),
            (By.NAME, "q"),
            (By.CSS_SELECTOR, 'input[type="text"]'),
            (By.CSS_SELECTOR, 'input[aria-label]'),
        ]

        for by, selector in selectors:
            try:
                search_box = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((by, selector))
                )
                log.info(f"Search box found via: {selector}")
                break
            except TimeoutException:
                log.warning(f"Selector not found: {selector}, trying next …")

        if search_box is None:
            log.error("Could not find search box. Saving screenshot for debug …")
            driver.save_screenshot("debug_screenshot.png")
            log.error("Screenshot saved as debug_screenshot.png — check what page opened.")
            return results

        search_box.clear()
        search_box.send_keys(query)
        time.sleep(1)
        search_box.send_keys(Keys.ENTER)
        time.sleep(4)

        # ── Step 3: Scroll the results panel to load more listings ────────
        log.info("Scrolling results panel to load listings …")
        results_panel_selector = 'div[role="feed"]'

        try:
            panel = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, results_panel_selector))
            )
        except TimeoutException:
            log.warning("Results panel not found — check if the search returned results.")
            return results

        # Scroll loop: keep scrolling until we have enough cards or hit the end
        last_count = 0
        stall_counter = 0

        while len(results) < max_results and stall_counter < 5:
            # Collect all currently loaded place cards
            cards = driver.find_elements(
                By.CSS_SELECTOR, 'a[href*="/maps/place/"]'
            )
            current_count = len(cards)

            log.info(f"  Loaded {current_count} listing cards so far …")

            # Scroll the panel down
            driver.execute_script("arguments[0].scrollTop += 1200;", panel)
            time.sleep(SCROLL_PAUSE)

            # Detect if new cards appeared; if not, increment stall counter
            if current_count == last_count:
                stall_counter += 1
            else:
                stall_counter = 0
                last_count = current_count

        # ── Step 4: Collect all place URLs first, then visit each directly ──
        # This completely avoids StaleElementReferenceException.
        # Strategy: grab hrefs as plain strings (not element refs), then
        # navigate to each URL independently — no DOM dependency at all.
        log.info("Collecting all place URLs from results panel …")

        place_urls = []
        seen = set()

        all_anchors = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
        for a in all_anchors:
            try:
                href = a.get_attribute("href")
                if href and href not in seen:
                    seen.add(href)
                    place_urls.append(href)
            except StaleElementReferenceException:
                continue

        place_urls = place_urls[:max_results]
        log.info(f"Collected {len(place_urls)} unique place URLs. Starting extraction …")

        for idx, place_url in enumerate(place_urls, start=1):
            restaurant = Restaurant()
            restaurant.maps_url = place_url

            try:
                # Navigate directly to the place page — completely avoids StaleElement
                driver.get(place_url)
                time.sleep(3)

                # ── Name ──────────────────────────────────────────────────
                try:
                    name_el = wait.until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, 'h1.DUwDvf, h1[class*="fontHeadlineLarge"]')
                        )
                    )
                    restaurant.name = name_el.text.strip()
                except TimeoutException:
                    restaurant.name = "N/A"

                # ── Rating ────────────────────────────────────────────────
                try:
                    rating_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'div.F7nice span[aria-hidden="true"]'
                    )
                    restaurant.rating = parse_rating(rating_el.text)
                except NoSuchElementException:
                    restaurant.rating = None

                # ── Review count ──────────────────────────────────────────
                try:
                    reviews_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'div.F7nice span[aria-label*="review"]'
                    )
                    restaurant.review_count = parse_review_count(
                        reviews_el.get_attribute("aria-label")
                    )
                except NoSuchElementException:
                    restaurant.review_count = 0

                # ── Cuisine / Category ────────────────────────────────────
                try:
                    cuisine_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'button.DkEaL, span.mgr77e'
                    )
                    restaurant.cuisine = cuisine_el.text.strip()
                except NoSuchElementException:
                    restaurant.cuisine = "N/A"

                # ── Address ───────────────────────────────────────────────
                try:
                    addr_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'button[data-item-id="address"] .Io6YTe'
                    )
                    restaurant.address = addr_el.text.strip()
                except NoSuchElementException:
                    restaurant.address = "N/A"

                # ── Phone ─────────────────────────────────────────────────
                try:
                    phone_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'button[data-item-id*="phone"] .Io6YTe'
                    )
                    restaurant.phone = phone_el.text.strip()
                except NoSuchElementException:
                    restaurant.phone = "N/A"

                # ── Website ───────────────────────────────────────────────
                try:
                    web_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'a[data-item-id="authority"] .Io6YTe'
                    )
                    restaurant.website = web_el.text.strip()
                except NoSuchElementException:
                    restaurant.website = "N/A"

                results.append(restaurant)
                log.info(
                    f"  [{idx}/{len(cards)}] ✓ {restaurant.name} "
                    f"| ⭐ {restaurant.rating} ({restaurant.review_count} reviews)"
                )

            except (StaleElementReferenceException, TimeoutException) as exc:
                log.warning(f"  [{idx}] Skipped — {type(exc).__name__}")
                continue

    finally:
        driver.quit()
        log.info("Browser closed.")

    return results


# ─────────────────────────────────────────────
#  SORTING
# ─────────────────────────────────────────────
def sort_results(data: list[Restaurant]) -> list[Restaurant]:
    """
    Sort restaurants by:
      1. Rating (descending)
      2. Review count (descending) — used as a tiebreaker
    Places with no rating are pushed to the bottom.
    """
    return sorted(
        data,
        key=lambda r: (r.rating is not None, r.rating or 0, r.review_count),
        reverse=True,
    )


# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────
def export_to_excel(data: list[Restaurant], filename: str) -> None:
    """
    Write scraped restaurant data to a formatted Excel workbook.
    Applies header styling, alternating row colors, and auto-fits columns.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NYC Restaurants"

    # ── Header row ────────────────────────────────────────────────────────
    headers = [
        "#", "Name", "Cuisine", "Rating", "Reviews",
        "Address", "Phone", "Website", "Google Maps URL"
    ]

    # Styling constants
    HEADER_BG    = "1A73E8"   # Google-blue
    HEADER_FONT  = "FFFFFF"
    ROW_ALT_BG   = "EAF1FB"   # Light blue for alternating rows
    BORDER_COLOR = "BFCFE8"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # Write and style header
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color=HEADER_FONT, size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, r in enumerate(data, start=2):
        row_data = [
            row_idx - 1,          # Row number
            r.name,
            r.cuisine,
            r.rating if r.rating is not None else "N/A",
            r.review_count,
            r.address,
            r.phone,
            r.website,
            r.maps_url,
        ]

        is_alt_row = (row_idx % 2 == 0)
        fill = PatternFill("solid", fgColor=ROW_ALT_BG) if is_alt_row else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = thin_border
            if fill:
                cell.fill = fill

            # Highlight top-rated restaurants (rating >= 4.5 and 500+ reviews)
            if col_idx == 4 and isinstance(value, float) and value >= 4.5:
                cell.font = Font(bold=True, color="0A6E0A")  # Dark green

        ws.row_dimensions[row_idx].height = 22

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = [5, 32, 20, 10, 12, 38, 16, 28, 50]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header row ─────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Summary row at the bottom ─────────────────────────────────────────
    summary_row = len(data) + 2
    ws.cell(row=summary_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(data)).font = Font(bold=True)
    avg_rating = (
        sum(r.rating for r in data if r.rating) / max(1, sum(1 for r in data if r.rating))
    )
    ws.cell(row=summary_row, column=4, value=round(avg_rating, 2)).font = Font(bold=True)

    wb.save(filename)
    log.info(f"Excel report saved → {filename}")


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
def main():
    log.info("=" * 55)
    log.info("  Google Maps Scraper — NYC Restaurants")
    log.info("=" * 55)

    # Step 1: Scrape data from Google Maps
    raw_data = scrape_google_maps(
        query=SEARCH_QUERY,
        max_results=MAX_RESULTS,
    )

    if not raw_data:
        log.error("No data collected. Check your internet connection or selectors.")
        return

    log.info(f"Collected {len(raw_data)} restaurants. Sorting …")

    # Step 2: Sort by rating + reviews
    sorted_data = sort_results(raw_data)

    # Step 3: Export to Excel
    export_to_excel(sorted_data, OUTPUT_FILE)

    log.info("=" * 55)
    log.info(f"  Done! Open '{OUTPUT_FILE}' to see your results.")
    log.info("=" * 55)


if __name__ == "__main__":
    main()
